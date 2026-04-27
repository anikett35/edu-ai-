"""
routers/teacher.py
──────────────────
Endpoints:

  POST /teacher/upload-material          – upload PDF/PPT/DOCX (teacher only)
  POST /teacher/create-quiz              – create quiz question (teacher only)
  GET  /teacher/materials                – list own uploaded materials (teacher only)
  GET  /teacher/materials/subject/{sub}  – list materials for a subject (all roles)
  POST /teacher/summarize/{material_id}  – summarize a document with Ollama (all roles)
"""

from datetime import datetime, timezone
from bson import ObjectId

from fastapi import APIRouter, Depends, UploadFile, File, Form, status, HTTPException
import httpx

from core.config   import settings
from core.security import require_teacher, get_current_user
from core.database import get_collection
from models.schemas import QuizQuestionCreate
from services.file_service import process_uploaded_file
from services.ml_service   import classify_difficulty

router = APIRouter(prefix="/teacher", tags=["Teacher"])


# ── Upload material ───────────────────────────────────────────────────────────

@router.post(
    "/upload-material",
    status_code=status.HTTP_201_CREATED,
    summary="Upload a PDF/PPT/DOCX study material (teacher only)",
)
async def upload_material(
    subject:  str        = Form(..., min_length=2, max_length=100),
    file:     UploadFile = File(...),
    teacher:  dict       = Depends(require_teacher),
):
    teacher_id   = teacher["user_id"]
    material_data = await process_uploaded_file(file=file, teacher_id=teacher_id, subject=subject)
    material_data["upload_date"] = datetime.now(timezone.utc)

    materials = get_collection("materials")
    result    = await materials.insert_one(material_data)

    return {
        "message":        "Material uploaded and processed successfully",
        "material_id":    str(result.inserted_id),
        "subject":        subject,
        "file_name":      material_data["file_name"],
        "chunks_created": material_data["chunk_count"],
    }


# ── Create quiz question ──────────────────────────────────────────────────────

@router.post(
    "/create-quiz",
    status_code=status.HTTP_201_CREATED,
    summary="Create a quiz question with auto difficulty classification",
)
async def create_quiz(
    body:    QuizQuestionCreate,
    teacher: dict = Depends(require_teacher),
):
    predicted_difficulty = classify_difficulty(body.question)

    quiz_doc = {
        "teacher_id":    teacher["user_id"],
        "subject":       body.subject,
        "question":      body.question,
        "options":       body.options,
        "correct_answer":body.correct_answer,
        "difficulty":    predicted_difficulty,
        "created_at":    datetime.now(timezone.utc),
    }

    quizzes = get_collection("quizzes")
    result  = await quizzes.insert_one(quiz_doc)

    return {
        "message":    "Quiz question created",
        "quiz_id":    str(result.inserted_id),
        "difficulty": predicted_difficulty,
    }


# ── List own materials (teacher only) ─────────────────────────────────────────

@router.get(
    "/materials",
    summary="List all materials uploaded by this teacher",
)
async def list_materials(teacher: dict = Depends(require_teacher)):
    materials = get_collection("materials")
    cursor    = materials.find(
        {"teacher_id": teacher["user_id"]},
        {"extracted_chunks": 0, "embeddings": 0},
    )
    docs = await cursor.to_list(length=100)

    return [
        {
            "id":          str(d["_id"]),
            "subject":     d["subject"],
            "file_name":   d["file_name"],
            "file_type":   d.get("file_type", "pdf"),
            "chunk_count": d.get("chunk_count", 0),
            "upload_date": d.get("upload_date"),
        }
        for d in docs
    ]


# ── List materials by subject (students + teachers) ───────────────────────────

@router.get(
    "/materials/subject/{subject}",
    summary="List all materials for a subject — accessible by both students and teachers",
)
async def materials_by_subject(
    subject:      str,
    current_user: dict = Depends(get_current_user),
):
    """
    Returns metadata only (no chunks/embeddings).
    Students use this to browse what their teacher has uploaded per subject.
    """
    materials = get_collection("materials")
    cursor    = materials.find(
        {"subject": subject},
        {"extracted_chunks": 0, "embeddings": 0},
    ).sort("upload_date", -1)
    docs = await cursor.to_list(length=50)

    return {
        "subject": subject,
        "total":   len(docs),
        "materials": [
            {
                "id":          str(d["_id"]),
                "file_name":   d["file_name"],
                "file_type":   d.get("file_type", "pdf"),
                "chunk_count": d.get("chunk_count", 0),
                "upload_date": d.get("upload_date"),
            }
            for d in docs
        ],
    }


# ── Summarize a material with Ollama ──────────────────────────────────────────

@router.post(
    "/summarize/{material_id}",
    summary="Summarize a specific uploaded document using Ollama (all roles)",
)
async def summarize_material(
    material_id:  str,
    current_user: dict = Depends(get_current_user),
):
    """
    Fetches the first N chunks of a document and asks Ollama to
    produce a concise structured summary with:
      - Overview (2-3 sentences)
      - Key Concepts (bullet list)
      - Important Topics
      - Study Tips

    Uses first 6 chunks (~2100 words) to stay within Mistral's context window.
    """
    # Fetch material
    materials = get_collection("materials")
    try:
        doc = await materials.find_one({"_id": ObjectId(material_id)})
    except Exception:
        raise HTTPException(status_code=400, detail="Invalid material ID")

    if not doc:
        raise HTTPException(status_code=404, detail="Material not found")

    chunks = doc.get("extracted_chunks", [])
    if not chunks:
        raise HTTPException(
            status_code=422,
            detail="No text content found in this material to summarize",
        )

    # Use first 6 chunks — ~2100 words, fits in Mistral 4096 token window
    sample_chunks = chunks[:6]
    content       = "\n\n---\n\n".join(sample_chunks)

    prompt = f"""You are an AI teaching assistant. A student has uploaded a study document.
Analyze the content below and produce a clear, structured summary.

=== DOCUMENT: {doc.get('file_name', 'Study Material')} ===
Subject: {doc.get('subject', 'Unknown')}

=== CONTENT SAMPLE ===
{content}

=== TASK ===
Provide a structured summary with these exact sections:

**OVERVIEW**
Write 2-3 sentences describing what this document covers.

**KEY CONCEPTS**
List the 5-8 most important concepts or terms found in this document.

**MAIN TOPICS**
List the main topics covered, with a one-line explanation each.

**STUDY TIPS**
Give 2-3 specific study tips based on the content of this document.

Keep the summary concise, educational, and student-friendly."""

    # Call Ollama
    ollama_timeout = httpx.Timeout(connect=10.0, read=180.0, write=30.0, pool=5.0)
    url            = f"{settings.ollama_base_url}/api/generate"

    try:
        async with httpx.AsyncClient(timeout=ollama_timeout) as client:
            response = await client.post(url, json={
                "model":  settings.ollama_model,
                "prompt": prompt,
                "stream": False,
                "options": {"temperature": 0.3, "num_predict": 800},
            })
            response.raise_for_status()
            summary_text = response.json().get("response", "").strip()

    except httpx.ConnectError:
        raise HTTPException(
            status_code=503,
            detail=f"Ollama not reachable at {settings.ollama_base_url}. Make sure 'ollama serve' is running.",
        )
    except httpx.ReadTimeout:
        raise HTTPException(
            status_code=503,
            detail="Ollama timed out generating the summary. Try again in 30 seconds.",
        )
    except Exception as e:
        raise HTTPException(status_code=503, detail=f"Ollama error: {str(e)}")

    return {
        "material_id": material_id,
        "file_name":   doc.get("file_name"),
        "subject":     doc.get("subject"),
        "chunk_count": len(chunks),
        "summary":     summary_text,
    }


# ── Download quiz template ────────────────────────────────────────────────────

@router.get(
    "/quiz-template",
    summary="Download Excel template for bulk quiz upload",
)
async def download_quiz_template(
    teacher: dict = Depends(require_teacher),
):
    """
    Returns an .xlsx file with:
      - Correct column headers
      - Data validation on Subject column (dropdown)
      - 3 example rows showing easy / medium / hard questions
      - Instructions sheet explaining the format
    """
    import io
    from fastapi.responses import StreamingResponse
    import openpyxl
    from openpyxl.styles import (
        PatternFill, Font, Alignment, Border, Side
    )
    from openpyxl.worksheet.datavalidation import DataValidation

    wb = openpyxl.Workbook()

    # ── Sheet 1: Quiz Data ────────────────────────────────────────────────────
    ws = wb.active
    ws.title = "Quiz Questions"

    # Column headers
    headers = ["subject", "question", "option_a", "option_b", "option_c", "option_d", "correct_answer"]
    col_widths = [20, 55, 22, 22, 22, 22, 18]

    # Header style
    header_fill  = PatternFill("solid", fgColor="312E81")   # deep indigo
    header_font  = Font(bold=True, color="FFFFFF", size=11, name="Calibri")
    center_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left_align   = Alignment(horizontal="left",   vertical="center", wrap_text=True)

    thin = Side(style="thin", color="C7D2FE")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for col_idx, (header, width) in enumerate(zip(headers, col_widths), start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.fill      = header_fill
        cell.font      = header_font
        cell.alignment = center_align
        cell.border    = border
        ws.column_dimensions[cell.column_letter].width = width

    ws.row_dimensions[1].height = 28

    # Subject dropdown validation
    subjects = "data_structures,machine_learning,dbms,operating_systems,full_stack"
    dv = DataValidation(
        type="list",
        formula1=f'"{subjects}"',
        allow_blank=False,
        showErrorMessage=True,
        errorTitle="Invalid Subject",
        error="Please select a valid subject from the dropdown.",
    )
    ws.add_data_validation(dv)
    dv.sqref = "A2:A1000"

    # Correct answer dropdown validation
    ans_dv = DataValidation(
        type="list",
        formula1='"option_a,option_b,option_c,option_d"',
        allow_blank=False,
        showErrorMessage=True,
        errorTitle="Invalid Answer",
        error="Must be one of: option_a, option_b, option_c, option_d",
    )
    ws.add_data_validation(ans_dv)
    ans_dv.sqref = "G2:G1000"

    # Example rows
    example_rows = [
        [
            "data_structures",
            "What is the time complexity of accessing an element in an array by index?",
            "O(n)", "O(1)", "O(log n)", "O(n²)", "option_b",
        ],
        [
            "machine_learning",
            "Explain the difference between overfitting and underfitting in machine learning models and describe techniques used to address each problem.",
            "Overfitting: high bias; Underfitting: high variance",
            "Overfitting: low training error, high test error; addressed by regularization, dropout, cross-validation",
            "Overfitting: high training error; Underfitting: low test error",
            "Both occur due to insufficient training data",
            "option_b",
        ],
        [
            "dbms",
            "What does ACID stand for in database transactions?",
            "Atomicity, Consistency, Isolation, Durability",
            "Availability, Consistency, Integrity, Distribution",
            "Atomicity, Concurrency, Isolation, Distribution",
            "Availability, Concurrency, Integrity, Durability",
            "option_a",
        ],
    ]

    row_fills = [
        PatternFill("solid", fgColor="EEF2FF"),   # easy   – light indigo
        PatternFill("solid", fgColor="FEF9C3"),   # medium – light yellow
        PatternFill("solid", fgColor="DCFCE7"),   # hard   – light green
    ]

    for row_idx, (row_data, fill) in enumerate(zip(example_rows, row_fills), start=2):
        for col_idx, value in enumerate(row_data, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.fill      = fill
            cell.border    = border
            cell.alignment = left_align if col_idx == 2 else center_align
            cell.font      = Font(size=10, name="Calibri")
        ws.row_dimensions[row_idx].height = 40

    # Freeze header row
    ws.freeze_panes = "A2"

    # ── Sheet 2: Instructions ─────────────────────────────────────────────────
    ws2 = wb.create_sheet("Instructions")
    ws2.column_dimensions["A"].width = 80

    instructions = [
        ("EduAI Quiz Upload Template — Instructions", "4338CA", True, 14),
        ("", None, False, 11),
        ("COLUMN REFERENCE", "1E1B4B", True, 11),
        ("subject        → One of: data_structures | machine_learning | dbms | operating_systems | full_stack", "374151", False, 10),
        ("question       → The question text. More complex questions = harder difficulty", "374151", False, 10),
        ("option_a/b/c/d → Four unique answer choices (all must be filled)", "374151", False, 10),
        ("correct_answer → Must be exactly: option_a, option_b, option_c, or option_d", "374151", False, 10),
        ("", None, False, 10),
        ("HOW DIFFICULTY IS CLASSIFIED (ML Model)", "1E1B4B", True, 11),
        ("Algorithm: Logistic Regression with TF-IDF vectorization (ngram_range=1-2)", "374151", False, 10),
        ("", None, False, 10),
        ("The model analyses your question text and classifies it as:", "374151", False, 10),
        ("  EASY   → Short questions, simple terms (e.g. 'What is X?', definitions)", "16A34A", False, 10),
        ("  MEDIUM → Multi-part questions, comparisons, explain-type questions", "D97706", False, 10),
        ("  HARD   → Complex analysis, derivations, 'compare A vs B with examples'", "DC2626", False, 10),
        ("", None, False, 10),
        ("TIPS FOR GOOD QUESTIONS", "1E1B4B", True, 11),
        ("✓ Keep question text clear and unambiguous", "374151", False, 10),
        ("✓ All 4 options must be unique (no duplicates)", "374151", False, 10),
        ("✓ One and only one correct answer", "374151", False, 10),
        ("✓ For Hard questions: use technical terminology, multi-step reasoning", "374151", False, 10),
        ("✓ Maximum rows per upload: 100 questions", "374151", False, 10),
        ("", None, False, 10),
        ("SUPPORTED FILE FORMATS", "1E1B4B", True, 11),
        (".xlsx (Excel) — Recommended", "374151", False, 10),
        (".csv (comma-separated) — Also accepted", "374151", False, 10),
    ]

    for row_idx, (text, color, bold, size) in enumerate(instructions, start=1):
        cell = ws2.cell(row=row_idx, column=1, value=text)
        if color:
            cell.font = Font(color=color, bold=bold, size=size, name="Calibri")
        cell.alignment = Alignment(horizontal="left", vertical="center")
        ws2.row_dimensions[row_idx].height = 18 if text else 8

    # Save to bytes
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": 'attachment; filename="EduAI_Quiz_Template.xlsx"'},
    )


# ── Bulk upload quiz from Excel/CSV ──────────────────────────────────────────

@router.post(
    "/quiz-bulk-upload",
    status_code=status.HTTP_201_CREATED,
    summary="Bulk upload quiz questions from Excel or CSV file",
)
async def bulk_upload_quiz(
    file:    UploadFile = File(...),
    teacher: dict       = Depends(require_teacher),
):
    """
    Accepts .xlsx or .csv files matching the EduAI quiz template.

    Processing pipeline:
      1. Validate file type
      2. Parse rows (openpyxl for xlsx, csv module for csv)
      3. Validate each row (subject, question, 4 options, correct_answer)
      4. Run ML difficulty classifier on each question text
      5. Bulk insert valid rows into MongoDB quizzes collection
      6. Return per-row results (success / error) + summary

    ML Classifier:
      Algorithm : Logistic Regression
      Features  : TF-IDF (unigrams + bigrams, max_features=5000, sublinear_tf=True)
      Classes   : easy | medium | hard
      Training  : 30 labelled CS questions (seed data) — retrained on new labels over time
    """
    import io
    import csv as csv_module

    suffix = file.filename.rsplit(".", 1)[-1].lower() if "." in file.filename else ""
    if suffix not in {"xlsx", "csv"}:
        raise HTTPException(
            status_code=status.HTTP_415_UNSUPPORTED_MEDIA_TYPE,
            detail="Only .xlsx and .csv files are accepted. Download the template from the Upload page.",
        )

    content = await file.read()
    if len(content) > 5 * 1024 * 1024:  # 5 MB
        raise HTTPException(status_code=413, detail="File too large (max 5 MB)")

    # ── Parse file ────────────────────────────────────────────────────────────
    REQUIRED_COLS = {"subject", "question", "option_a", "option_b", "option_c", "option_d", "correct_answer"}
    VALID_SUBJECTS = {"data_structures", "machine_learning", "dbms", "operating_systems", "full_stack"}
    VALID_ANSWERS  = {"option_a", "option_b", "option_c", "option_d"}

    raw_rows = []

    if suffix == "xlsx":
        import openpyxl
        wb  = openpyxl.load_workbook(io.BytesIO(content), read_only=True, data_only=True)
        ws  = wb.active
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            raise HTTPException(status_code=422, detail="Excel file is empty")
        headers = [str(h).strip().lower() if h else "" for h in rows[0]]
        for row in rows[1:]:
            if all(cell is None for cell in row):
                continue   # skip blank rows
            raw_rows.append(dict(zip(headers, [str(c).strip() if c is not None else "" for c in row])))
        wb.close()

    else:  # csv
        text   = content.decode("utf-8-sig")   # handle BOM
        reader = csv_module.DictReader(io.StringIO(text))
        reader.fieldnames = [f.strip().lower() for f in (reader.fieldnames or [])]
        raw_rows = [{k.strip().lower(): v.strip() for k, v in row.items()} for row in reader]

    if not raw_rows:
        raise HTTPException(status_code=422, detail="No data rows found in file")
    if len(raw_rows) > 100:
        raise HTTPException(status_code=422, detail="Maximum 100 questions per upload")

    # Check headers
    file_cols = set(raw_rows[0].keys())
    missing   = REQUIRED_COLS - file_cols
    if missing:
        raise HTTPException(
            status_code=422,
            detail=f"Missing required columns: {', '.join(sorted(missing))}. Download the template for correct format.",
        )

    # ── Validate + classify ───────────────────────────────────────────────────
    quizzes_col = get_collection("quizzes")
    from datetime import datetime, timezone
    from services.ml_service import classify_difficulty

    results      = []
    to_insert    = []
    success_count = 0
    error_count   = 0

    for row_num, row in enumerate(raw_rows, start=2):  # start=2 because row 1 is header
        subject        = row.get("subject", "").strip().lower()
        question       = row.get("question", "").strip()
        option_a       = row.get("option_a", "").strip()
        option_b       = row.get("option_b", "").strip()
        option_c       = row.get("option_c", "").strip()
        option_d       = row.get("option_d", "").strip()
        correct_answer_key = row.get("correct_answer", "").strip().lower()

        # Per-row validation
        row_errors = []
        if subject not in VALID_SUBJECTS:
            row_errors.append(f"Invalid subject '{subject}'")
        if len(question) < 10:
            row_errors.append("Question too short (min 10 chars)")
        options = [option_a, option_b, option_c, option_d]
        if any(o == "" for o in options):
            row_errors.append("All 4 options must be filled")
        if len(set(options)) != 4:
            row_errors.append("All 4 options must be unique")
        if correct_answer_key not in VALID_ANSWERS:
            row_errors.append(f"correct_answer must be option_a/b/c/d, got '{correct_answer_key}'")

        if row_errors:
            error_count += 1
            results.append({
                "row": row_num, "status": "error",
                "question": question[:60] + ("…" if len(question) > 60 else ""),
                "errors": row_errors,
            })
            continue

        # Map option key → actual text
        option_map    = {"option_a": option_a, "option_b": option_b,
                         "option_c": option_c, "option_d": option_d}
        correct_text  = option_map[correct_answer_key]

        # ML difficulty classification
        try:
            difficulty = classify_difficulty(question)
        except Exception:
            difficulty = "medium"  # safe fallback

        to_insert.append({
            "teacher_id":    teacher["user_id"],
            "subject":       subject,
            "question":      question,
            "options":       options,
            "correct_answer": correct_text,
            "difficulty":    difficulty,
            "created_at":    datetime.now(timezone.utc),
            "source":        "bulk_upload",
            "original_file": file.filename,
        })
        success_count += 1
        results.append({
            "row":        row_num,
            "status":     "success",
            "question":   question[:60] + ("…" if len(question) > 60 else ""),
            "subject":    subject,
            "difficulty": difficulty,
        })

    # Bulk insert
    if to_insert:
        await quizzes_col.insert_many(to_insert)

    # Difficulty breakdown
    difficulties = [r["difficulty"] for r in results if r["status"] == "success"]
    breakdown = {
        "easy":   difficulties.count("easy"),
        "medium": difficulties.count("medium"),
        "hard":   difficulties.count("hard"),
    }

    return {
        "message":       f"Processed {len(raw_rows)} rows: {success_count} created, {error_count} failed",
        "total_rows":    len(raw_rows),
        "success_count": success_count,
        "error_count":   error_count,
        "difficulty_breakdown": breakdown,
        "ml_classifier": "Logistic Regression + TF-IDF (ngram 1-2, max_features=5000, sublinear_tf=True)",
        "results":       results,
    }